#!/usr/bin/env python3
"""
수료증 대량 생성기 (Certificate Bulk Generator)
- 구글 시트 또는 엑셀 데이터 + 템플릿 이미지 → 수료증 PNG 자동 생성
"""

import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path

from PIL import Image, ImageDraw, ImageFont


def load_config(config_path: str) -> dict:
    """config.json 로드"""
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


def format_date_korean(dt) -> str:
    """datetime 또는 문자열을 '2026년 3월 14일' 형식으로 변환"""
    if isinstance(dt, datetime):
        return f"{dt.year}년 {dt.month}월 {dt.day}일"
    if isinstance(dt, str):
        dt = dt.strip()
        # 'YYYY-MM-DD' 형식 파싱
        try:
            parsed = datetime.strptime(dt, "%Y-%m-%d")
            return f"{parsed.year}년 {parsed.month}월 {parsed.day}일"
        except ValueError:
            pass
        # 'YYYY. M. D.' 또는 'YYYY. M. D' 형식 (구글 시트 날짜)
        try:
            cleaned = dt.replace(".", "").strip()
            parts = cleaned.split()
            if len(parts) == 3:
                y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
                return f"{y}년 {m}월 {d}일"
        except (ValueError, IndexError):
            pass
        # 이미 한국어 형식이면 그대로 반환
        return dt
    return str(dt)


def read_google_sheet(sheet_url: str, credentials_path: str) -> list[dict]:
    """구글 시트에서 수료자 데이터 읽기"""
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        print("[오류] gspread 또는 google-auth 라이브러리가 설치되지 않았습니다.")
        print("  pip install gspread 를 실행해주세요.")
        sys.exit(1)

    # 인증
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets.readonly",
    ]
    creds = Credentials.from_service_account_file(credentials_path, scopes=scopes)
    gc = gspread.authorize(creds)

    # 시트 열기 (URL 또는 시트 ID 지원)
    if sheet_url.startswith("http"):
        sh = gc.open_by_url(sheet_url)
    else:
        sh = gc.open_by_key(sheet_url)

    ws = sh.sheet1
    all_values = ws.get_all_values()

    if not all_values:
        print("[오류] 구글 시트가 비어있습니다.")
        sys.exit(1)

    # 헤더
    headers = [h.strip() for h in all_values[0]]

    # 필수 열 확인
    required = {"발급번호", "성명", "기간_시작", "기간_종료"}
    found = set(headers)
    missing = required - found
    if missing:
        print(f"[오류] 구글 시트에 필수 열이 없습니다: {', '.join(missing)}")
        print(f"  현재 열: {headers}")
        print(f"  필수 열: 발급번호, 성명, 기간_시작, 기간_종료")
        sys.exit(1)

    # 데이터 읽기
    records = []
    for row_idx, row in enumerate(all_values[1:], start=2):
        record = {}
        for i, value in enumerate(row):
            if i < len(headers) and headers[i]:
                record[headers[i]] = value.strip() if value else ""

        # 빈 행 스킵
        if not record.get("성명"):
            print(f"  [경고] {row_idx}행: 성명이 비어있어 건너뜁니다.")
            continue

        records.append(record)

    return records


def read_excel(filepath: str) -> list[dict]:
    """엑셀 파일에서 수료자 데이터 읽기"""
    from openpyxl import load_workbook

    wb = load_workbook(filepath, read_only=True)
    ws = wb.active

    # 헤더 읽기
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(str(cell.value).strip() if cell.value else "")

    # 필수 열 확인
    required = {"발급번호", "성명", "기간_시작", "기간_종료"}
    found = set(headers)
    missing = required - found
    if missing:
        print(f"[오류] 엑셀에 필수 열이 없습니다: {', '.join(missing)}")
        print(f"  현재 열: {headers}")
        print(f"  필수 열: 발급번호, 성명, 기간_시작, 기간_종료")
        sys.exit(1)

    # 데이터 읽기
    records = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        record = {}
        for i, cell in enumerate(row):
            if i < len(headers) and headers[i]:
                record[headers[i]] = cell.value

        # 빈 행 스킵
        if not record.get("성명"):
            print(f"  [경고] {row_idx}행: 성명이 비어있어 건너뜁니다.")
            continue

        records.append(record)

    wb.close()
    return records


def validate_records(records: list[dict]) -> list[dict]:
    """데이터 유효성 검사"""
    valid = []
    issue_numbers = set()

    for i, rec in enumerate(records):
        issues = []

        if not rec.get("발급번호"):
            issues.append("발급번호 누락")
        elif rec["발급번호"] in issue_numbers:
            print(f"  [경고] 발급번호 '{rec['발급번호']}' 중복 (덮어쓰기됩니다)")

        issue_numbers.add(rec.get("발급번호", ""))

        if not rec.get("기간_시작"):
            issues.append("기간_시작 누락")
        if not rec.get("기간_종료"):
            issues.append("기간_종료 누락")

        if issues:
            print(f"  [경고] {i+2}행 ({rec.get('성명', '?')}): {', '.join(issues)} → 건너뜁니다.")
        else:
            valid.append(rec)

    return valid


def generate_certificate(
    template_img: Image.Image,
    record: dict,
    config: dict,
    base_dir: str,
    font_cache: dict = None,
) -> Image.Image:
    """단일 수료증 이미지 생성"""
    img = template_img.copy()
    draw = ImageDraw.Draw(img)

    fonts_config = config["fonts"]

    # 폰트 로드 캐시 (외부에서 전달받거나 새로 생성)
    if font_cache is None:
        font_cache = {}
        for key, fc in fonts_config.items():
            font_path = os.path.join(base_dir, fc["path"])
            font_cache[key] = ImageFont.truetype(font_path, fc["size"])

    # 1) 고정 텍스트 (라벨 + 정적 값) 렌더링
    for st in config.get("static_texts", []):
        fc = fonts_config[st["font_key"]]
        draw.text(
            (st["x"], st["y"]),
            st["text"],
            font=font_cache[st["font_key"]],
            fill=tuple(fc["color"]),
        )

    # 2) 동적 필드 렌더링
    fields_config = config.get("dynamic_fields", config.get("fields", {}))

    # 발급번호
    if "발급번호" in fields_config:
        field = fields_config["발급번호"]
        fc = fonts_config[field["font_key"]]
        draw.text(
            (field["x"], field["y"]),
            str(record["발급번호"]),
            font=font_cache[field["font_key"]],
            fill=tuple(fc["color"]),
        )

    # 성명
    if "성명" in fields_config:
        field = fields_config["성명"]
        fc = fonts_config[field["font_key"]]
        draw.text(
            (field["x"], field["y"]),
            str(record["성명"]),
            font=font_cache[field["font_key"]],
            fill=tuple(fc["color"]),
        )

    # 기간
    if "기간" in fields_config:
        start_str = format_date_korean(record["기간_시작"])
        end_str = format_date_korean(record["기간_종료"])
        period_text = f"{start_str} ~ {end_str}"

        field = fields_config["기간"]
        fc = fonts_config[field["font_key"]]
        draw.text(
            (field["x"], field["y"]),
            period_text,
            font=font_cache[field["font_key"]],
            fill=tuple(fc["color"]),
        )

    return img


def generate_calibration(config: dict, base_dir: str):
    """좌표 보정용 이미지 생성 — 각 필드 위치에 빨간 십자 마커 표시"""
    template_path = os.path.join(base_dir, config["template"]["image_path"])
    img = Image.open(template_path).copy()
    draw = ImageDraw.Draw(img)

    fields_config = config["fields"]
    marker_size = 40

    for field_name, field in fields_config.items():
        x, y = field["x"], field["y"]
        # 십자 마커
        draw.line([(x - marker_size, y), (x + marker_size, y)], fill=(255, 0, 0), width=3)
        draw.line([(x, y - marker_size), (x, y + marker_size)], fill=(255, 0, 0), width=3)
        # 필드 이름 레이블
        draw.text((x + marker_size + 5, y - 15), field_name, fill=(255, 0, 0))

    output_path = os.path.join(base_dir, "templates", "_calibrate.png")
    img.save(output_path)
    print(f"[캘리브레이션] 저장 완료: {output_path}")
    print("  빨간 십자 마커 위치를 확인하고 config.json의 좌표를 조정하세요.")


def find_credentials(base_dir: str) -> str:
    """프로젝트 폴더에서 서비스 계정 JSON 키 파일 자동 탐색"""
    for f in os.listdir(base_dir):
        if f.endswith(".json") and f != "config.json":
            filepath = os.path.join(base_dir, f)
            try:
                with open(filepath, "r") as fh:
                    data = json.load(fh)
                    if "type" in data and data["type"] == "service_account":
                        return filepath
            except (json.JSONDecodeError, KeyError):
                continue
    return ""


def main():
    parser = argparse.ArgumentParser(
        description="수료증 대량 생성기",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
사용 예시:
  # 구글 시트에서 데이터 읽기 (기본)
  python generate.py --sheet "https://docs.google.com/spreadsheets/d/..."

  # 엑셀 파일에서 데이터 읽기
  python generate.py --data data/participants.xlsx

  # 출력 폴더 지정
  python generate.py --sheet "URL" --output ./2026년_수료증

  # 특정 행만 생성 (테스트용)
  python generate.py --sheet "URL" --rows 1-3

  # 좌표 보정 모드
  python generate.py --calibrate
        """,
    )
    parser.add_argument("--sheet", type=str, help="구글 시트 URL 또는 시트 ID")
    parser.add_argument("--data", type=str, help="수료자 명단 엑셀 파일 경로 (구글 시트 대신 사용)")
    parser.add_argument("--credentials", type=str, help="서비스 계정 JSON 키 파일 경로 (미지정 시 자동 탐색)")
    parser.add_argument("--config", type=str, default="config.json", help="설정 파일 경로 (기본: config.json)")
    parser.add_argument("--output", type=str, help="출력 폴더 (기본: config의 output_dir)")
    parser.add_argument("--rows", type=str, help="생성할 행 범위 (예: 1-3, 5)")
    parser.add_argument("--calibrate", action="store_true", help="좌표 보정 모드")

    args = parser.parse_args()

    # 기준 디렉토리 (스크립트 위치)
    base_dir = os.path.dirname(os.path.abspath(__file__))

    # 설정 로드
    config_path = os.path.join(base_dir, args.config)
    if not os.path.exists(config_path):
        print(f"[오류] 설정 파일을 찾을 수 없습니다: {config_path}")
        sys.exit(1)

    config = load_config(config_path)

    # 캘리브레이션 모드
    if args.calibrate:
        generate_calibration(config, base_dir)
        return

    # 데이터 소스 확인
    if not args.sheet and not args.data:
        print("[오류] --sheet 또는 --data 옵션으로 데이터 소스를 지정해주세요.")
        print("  예: python generate.py --sheet \"https://docs.google.com/spreadsheets/d/...\"")
        print("  예: python generate.py --data data/participants.xlsx")
        sys.exit(1)

    # 출력 폴더
    output_dir = args.output or os.path.join(base_dir, config["template"]["output_dir"])
    os.makedirs(output_dir, exist_ok=True)

    # 폰트 파일 존재 확인
    for font_key, fc in config["fonts"].items():
        font_path = os.path.join(base_dir, fc["path"])
        if not os.path.exists(font_path):
            print(f"[오류] 폰트 파일을 찾을 수 없습니다: {font_path}")
            sys.exit(1)

    # 템플릿 이미지 로드
    template_path = os.path.join(base_dir, config["template"]["image_path"])
    if not os.path.exists(template_path):
        print(f"[오류] 템플릿 이미지를 찾을 수 없습니다: {template_path}")
        sys.exit(1)

    template_img = Image.open(template_path)
    print(f"[템플릿] {template_path} ({template_img.size[0]}x{template_img.size[1]})")

    # 데이터 읽기
    if args.sheet:
        # 구글 시트 모드
        creds_path = args.credentials
        if not creds_path:
            creds_path = find_credentials(base_dir)
        if not creds_path:
            print("[오류] 서비스 계정 JSON 키 파일을 찾을 수 없습니다.")
            print("  --credentials 옵션으로 경로를 지정하거나, 프로젝트 폴더에 JSON 키 파일을 넣어주세요.")
            sys.exit(1)

        print(f"[구글 시트] {args.sheet}")
        print(f"[인증] {os.path.basename(creds_path)}")
        records = read_google_sheet(args.sheet, creds_path)
    else:
        # 엑셀 모드
        data_path = args.data
        if not os.path.isabs(data_path):
            data_path = os.path.join(base_dir, data_path)

        if not os.path.exists(data_path):
            print(f"[오류] 엑셀 파일을 찾을 수 없습니다: {data_path}")
            sys.exit(1)

        print(f"[데이터] {data_path} 읽는 중...")
        records = read_excel(data_path)

    print(f"  {len(records)}건 로드됨")

    # 유효성 검사
    records = validate_records(records)
    print(f"  {len(records)}건 유효")

    if not records:
        print("[완료] 생성할 데이터가 없습니다.")
        return

    # 행 범위 필터
    if args.rows:
        if "-" in args.rows:
            start, end = args.rows.split("-")
            start, end = int(start) - 1, int(end)
            records = records[start:end]
        else:
            idx = int(args.rows) - 1
            records = [records[idx]] if idx < len(records) else []
        print(f"  → {len(records)}건 선택됨 (--rows {args.rows})")

    # 생성
    success = 0
    fail = 0
    filename_template = config["template"].get("output_filename", "{발급번호}_{성명}")
    output_format = config["template"].get("output_format", "png")

    print(f"\n[생성 시작] {len(records)}건")
    print("-" * 50)

    for i, record in enumerate(records, 1):
        try:
            cert_img = generate_certificate(template_img, record, config, base_dir)

            # 파일명 생성
            filename = filename_template.format(
                발급번호=record["발급번호"],
                성명=record["성명"],
            )
            # 파일명에서 위험한 문자 제거
            filename = "".join(c for c in filename if c not in r'\/:*?"<>|')
            output_path = os.path.join(output_dir, f"{filename}.{output_format}")

            cert_img.save(output_path, quality=95)
            print(f"  [{i}/{len(records)}] {record['성명']} → {os.path.basename(output_path)}")
            success += 1

        except Exception as e:
            print(f"  [{i}/{len(records)}] {record.get('성명', '?')} → 실패: {e}")
            fail += 1

    print("-" * 50)
    print(f"[완료] 성공: {success}건, 실패: {fail}건")
    print(f"  출력 폴더: {output_dir}")


if __name__ == "__main__":
    main()
